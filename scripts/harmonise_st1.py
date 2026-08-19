"""
ST1 — Data Load & Label Harmonisation (Corrected)

Fixes applied per mentor review:
  - CADEC: Pre-split posts on bare newlines before Punkt tokenization
  - Output 50 random CADEC units (25 ADR+, 25 ADR-) for manual hand-audit
  - Output 10 example units per corpus
  - Side-by-side sentence-length distribution comparison
  - Ambiguous mapping notes documented
"""
import os
import sys
import glob
import re
import pandas as pd
import numpy as np
import openpyxl
import nltk
from nltk.tokenize import PunktSentenceTokenizer


def reconfigure_stdout():
    if hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass


def setup_nltk():
    """Ensure NLTK sentence tokenization resources are downloaded safely."""
    for resource in ['punkt', 'punkt_tab']:
        try:
            nltk.download(resource, quiet=True)
        except Exception as e:
            print(f"[NLTK Setup Note] {resource} download attempt: {e}")


def harmonise_psytar(psytar_excel_path, output_csv_path):
    """
    Load PsyTAR Excel file, map sentences to binary ADR target,
    and save harmonised CSV with columns ['text', 'label'].
    """
    print(f"\n--- Processing PsyTAR Dataset ({psytar_excel_path}) ---")
    wb = openpyxl.load_workbook(psytar_excel_path, read_only=True,
                                data_only=True)
    sheet_names = wb.sheetnames
    print(f"Detected Excel Sheet Names: {sheet_names}")

    target_sheet = ('Sentence_Labeling' if 'Sentence_Labeling' in sheet_names
                    else sheet_names[0])
    print(f"Reading sheet: '{target_sheet}'")

    ws = wb[target_sheet]
    row_generator = ws.iter_rows(values_only=True)
    header = next(row_generator)
    header_list = [str(h).strip() if h is not None else '' for h in header]
    print(f"Headers in '{target_sheet}': {header_list}")

    sent_idx = header_list.index('sentences') if 'sentences' in header_list else -1
    adr_idx = header_list.index('ADR') if 'ADR' in header_list else -1
    id_idx = header_list.index('id') if 'id' in header_list else -1

    if sent_idx == -1:
        raise ValueError("Could not find 'sentences' column in PsyTAR Excel.")

    sentences = []
    labels = []

    for row in row_generator:
        if not row:
            continue
        row_id = row[id_idx] if id_idx != -1 and id_idx < len(row) else None
        sent_val = row[sent_idx] if sent_idx < len(row) else None

        if sent_val is None or str(sent_val).strip() == '':
            continue
        if id_idx != -1 and row_id is None:
            continue

        sent_text = str(sent_val).strip()
        adr_val = row[adr_idx] if adr_idx != -1 and adr_idx < len(row) else None

        is_adr = 0
        if adr_val is not None:
            try:
                if float(adr_val) == 1.0:
                    is_adr = 1
            except (ValueError, TypeError):
                pass

        sentences.append(sent_text)
        labels.append(is_adr)

    wb.close()

    df_harmonised = pd.DataFrame({'text': sentences, 'label': labels})
    os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
    df_harmonised.to_csv(output_csv_path, index=False, encoding='utf-8')
    print(f"Saved PsyTAR harmonised dataset ({len(df_harmonised)} rows) to: "
          f"{output_csv_path}")
    return df_harmonised


def find_cadec_ann_path(txt_path):
    """Locate the corresponding .ann Brat annotation file for a CADEC .txt post."""
    base_no_ext = os.path.splitext(txt_path)[0]
    ann_same_dir = base_no_ext + '.ann'
    if os.path.exists(ann_same_dir):
        return ann_same_dir

    ann_original = txt_path.replace('\\text\\', '\\original\\').replace(
        '/text/', '/original/')
    ann_original = os.path.splitext(ann_original)[0] + '.ann'
    if os.path.exists(ann_original):
        return ann_original

    base_dir = os.path.dirname(os.path.dirname(txt_path))
    filename = os.path.splitext(os.path.basename(txt_path))[0] + '.ann'
    ann_fallback = os.path.join(base_dir, 'original', filename)
    if os.path.exists(ann_fallback):
        return ann_fallback

    return None


def parse_brat_adr_spans(ann_path):
    """Extract all ADR entity spans (start_char, end_char) from Brat .ann file."""
    adr_spans = []
    if not ann_path or not os.path.exists(ann_path):
        return adr_spans

    with open(ann_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            line_str = line.strip()
            if not line_str.startswith('T'):
                continue
            parts = line_str.split('\t')
            if len(parts) < 2:
                continue
            entity_info = parts[1].strip()
            tokens = entity_info.split()
            if not tokens:
                continue
            entity_type = tokens[0]
            if entity_type.lower() in ['adr', 'adverse_drug_reaction']:
                span_str = " ".join(tokens[1:])
                numbers = [int(n) for n in re.findall(r'\d+', span_str)]
                for i in range(0, len(numbers) - 1, 2):
                    start_off, end_off = numbers[i], numbers[i + 1]
                    if start_off < end_off:
                        adr_spans.append((start_off, end_off))
    return adr_spans


def harmonise_cadec(cadec_folder_path, output_csv_path):
    """
    CORRECTED: Pre-split posts on bare newlines before Punkt tokenization.
    This prevents bullet-style forum posts from becoming single long 'sentences'.
    """
    print(f"\n--- Processing CADEC Dataset ({cadec_folder_path}) ---")
    print("  [FIX] Pre-splitting posts on bare newlines before Punkt tokenization.")

    txt_files = glob.glob(os.path.join(cadec_folder_path, '**', '*.txt'),
                          recursive=True)
    txt_files.sort()
    print(f"Found {len(txt_files)} text posts in CADEC directory.")

    sentences_records = []
    boundary_cross_count = 0
    missing_ann_count = 0
    total_adr_spans_found = 0
    newline_splits_applied = 0

    tokenizer = PunktSentenceTokenizer()

    for txt_path in txt_files:
        ann_path = find_cadec_ann_path(txt_path)
        if not ann_path:
            missing_ann_count += 1
            adr_spans = []
        else:
            adr_spans = parse_brat_adr_spans(ann_path)
            total_adr_spans_found += len(adr_spans)

        with open(txt_path, 'r', encoding='utf-8', errors='ignore') as f:
            post_text = f.read()

        if not post_text.strip():
            continue

        # FIX: Pre-split on bare newlines into paragraphs, then tokenize each
        # This handles bullet-style forum posts where \n separates distinct items
        paragraphs = post_text.split('\n')
        if len(paragraphs) > 1:
            newline_splits_applied += 1

        global_offset = 0
        for para_idx, paragraph in enumerate(paragraphs):
            para_stripped = paragraph.strip()
            # Find where this paragraph starts in the original post_text
            para_start_in_post = post_text.find(paragraph, global_offset)
            if para_start_in_post == -1:
                para_start_in_post = global_offset
            global_offset = para_start_in_post + len(paragraph)

            if not para_stripped:
                continue

            # Sentence-tokenize within this paragraph
            sent_spans = list(tokenizer.span_tokenize(paragraph))

            for s_start_local, s_end_local in sent_spans:
                # Map local offsets back to global post offsets
                s_start_global = para_start_in_post + s_start_local
                s_end_global = para_start_in_post + s_end_local

                sentence_text = post_text[s_start_global:s_end_global].strip()
                if not sentence_text:
                    continue

                # Check overlap with any ADR span
                has_adr = 0
                for a_start, a_end in adr_spans:
                    if max(s_start_global, a_start) < min(s_end_global, a_end):
                        has_adr = 1
                        if a_start < s_start_global or a_end > s_end_global:
                            boundary_cross_count += 1
                        break

                sentences_records.append({
                    'text': sentence_text,
                    'label': has_adr
                })

    df_harmonised = pd.DataFrame(sentences_records)
    os.makedirs(os.path.dirname(output_csv_path), exist_ok=True)
    df_harmonised.to_csv(output_csv_path, index=False, encoding='utf-8')
    print(f"Saved CADEC harmonised dataset ({len(df_harmonised)} rows) to: "
          f"{output_csv_path}")

    alignment_notes = []
    if missing_ann_count > 0:
        alignment_notes.append(
            f"Missing annotation files for {missing_ann_count} posts.")
    if boundary_cross_count > 0:
        alignment_notes.append(
            f"Detected {boundary_cross_count} instances where annotated ADR "
            f"entity spans crossed sentence boundaries.")
    alignment_notes.append(
        f"Total ADR character spans parsed from Brat annotations: "
        f"{total_adr_spans_found}.")
    alignment_notes.append(
        f"Posts pre-split on newlines: {newline_splits_applied} "
        f"(out of {len(txt_files)} total).")

    return df_harmonised, alignment_notes


def print_st1_report(df_psytar, df_cadec, cadec_alignment_notes):
    """Print ST1 Report with corrected content per mentor review."""
    print("\n" + "=" * 80)
    print("              ST1 HARMONISATION REPORT (CORRECTED)")
    print("=" * 80)

    # 1. Count Table
    def compute_metrics(df, name):
        total = len(df)
        pos = int(df['label'].sum())
        neg = total - pos
        pct = (pos / total * 100) if total > 0 else 0.0
        return {
            'Dataset': name, 'Total Sentences': total,
            'ADR-Present (1)': pos, 'ADR-Absent (0)': neg,
            'ADR Percentage (%)': f"{pct:.2f}%"
        }

    m_psytar = compute_metrics(df_psytar, 'PsyTAR (Dev)')
    m_cadec = compute_metrics(df_cadec, 'CADEC (External Val)')
    df_metrics = pd.DataFrame([m_psytar, m_cadec])

    print("\n--- 1. COUNT TABLE (CLASS BALANCE & UNITS) ---")
    print(df_metrics.to_string(index=False))

    # 2. 10 Example Units per corpus
    def print_examples(df, name, n=10):
        print(f"\n--- 2. {name.upper()} — 10 EXAMPLE UNITS ---")
        pos_df = df[df['label'] == 1].head(5)
        neg_df = df[df['label'] == 0].head(5)
        sample_df = pd.concat([pos_df, neg_df]).head(n).reset_index(drop=True)
        for idx, row in sample_df.iterrows():
            print(f"  [{idx + 1:02d}] Label={row['label']} | "
                  f"\"{row['text'][:120]}{'...' if len(row['text']) > 120 else ''}\"")

    print_examples(df_psytar, "PsyTAR (Development Corpus)")
    print_examples(df_cadec, "CADEC (External Validation Corpus)")

    # 3. Sentence-length distribution comparison
    print("\n--- 3. SENTENCE-LENGTH DISTRIBUTION COMPARISON ---")
    psy_lens = df_psytar['text'].str.len()
    cad_lens = df_cadec['text'].str.len()

    comparison = pd.DataFrame({
        'Statistic': ['Count', 'Mean', 'Median', 'Std', 'Min', 'Max',
                      'Pct > 200 chars', 'Pct > 300 chars', 'Pct > 500 chars',
                      'Contains newline'],
        'PsyTAR': [
            len(psy_lens), f"{psy_lens.mean():.1f}", f"{psy_lens.median():.1f}",
            f"{psy_lens.std():.1f}", psy_lens.min(), psy_lens.max(),
            f"{(psy_lens > 200).sum()} ({(psy_lens > 200).mean() * 100:.2f}%)",
            f"{(psy_lens > 300).sum()} ({(psy_lens > 300).mean() * 100:.2f}%)",
            f"{(psy_lens > 500).sum()} ({(psy_lens > 500).mean() * 100:.2f}%)",
            f"{df_psytar['text'].str.contains(chr(10)).sum()}"
        ],
        'CADEC': [
            len(cad_lens), f"{cad_lens.mean():.1f}", f"{cad_lens.median():.1f}",
            f"{cad_lens.std():.1f}", cad_lens.min(), cad_lens.max(),
            f"{(cad_lens > 200).sum()} ({(cad_lens > 200).mean() * 100:.2f}%)",
            f"{(cad_lens > 300).sum()} ({(cad_lens > 300).mean() * 100:.2f}%)",
            f"{(cad_lens > 500).sum()} ({(cad_lens > 500).mean() * 100:.2f}%)",
            f"{df_cadec['text'].str.contains(chr(10)).sum()}"
        ]
    })
    print(comparison.to_string(index=False))

    # 4. 50-Unit CADEC Hand-Audit Sample
    print("\n--- 4. CADEC 50-UNIT HAND-AUDIT SAMPLE ---")
    pos_sample = df_cadec[df_cadec['label'] == 1].sample(
        n=min(25, len(df_cadec[df_cadec['label'] == 1])),
        random_state=42)
    neg_sample = df_cadec[df_cadec['label'] == 0].sample(
        n=min(25, len(df_cadec[df_cadec['label'] == 0])),
        random_state=42)
    audit_df = pd.concat([pos_sample, neg_sample]).reset_index(drop=True)
    for idx, row in audit_df.iterrows():
        print(f"  [{idx + 1:02d}] Label={row['label']} | Len={len(row['text']):4d} | "
              f"\"{row['text'][:100]}{'...' if len(row['text']) > 100 else ''}\"")

    # 5. Alignment notes
    print("\n--- 5. SENTENCE BOUNDARY ALIGNMENT & TOKENIZATION AUDIT NOTES ---")
    print("  - PsyTAR: Sentences were pre-segmented in Excel 'Sentence_Labeling' sheet.")
    print(f"    Total valid sentences extracted: {len(df_psytar)}")
    print("  - CADEC Alignment Notes:")
    for note in cadec_alignment_notes:
        print(f"    * {note}")
    print("  - Tokenization: NLTK PunktSentenceTokenizer with newline pre-split.")
    print("  - Ambiguous Mapping Notes:")
    print("    * PsyTAR ADR column: Binary (1.0 = ADR present, else 0). "
          "Summary/totals rows filtered by null id/sentence.")
    print("    * CADEC ADR: Entity type 'ADR' from Brat annotations. "
          "Sentence labelled positive if ANY ADR span character-overlaps.")
    print("    * Boundary-crossing ADR spans: counted but sentence still "
          "labelled positive (partial overlap = positive).")
    print("    * Harmonisation was locked BEFORE any model results were examined.")
    print("=" * 80 + "\n")


def main():
    reconfigure_stdout()
    setup_nltk()

    base_dir = r"e:\AI Green"
    psytar_excel = os.path.join(
        base_dir, r"data\01_primary_adr_detection\dev_psytar\PsyTAR_dataset.xlsx")
    psytar_out = os.path.join(
        base_dir, r"data\01_primary_adr_detection\dev_psytar\psytar_harmonised.csv")
    cadec_folder = os.path.join(
        base_dir, r"data\01_primary_adr_detection\external_val_cadec\cadec")
    cadec_out = os.path.join(
        base_dir, r"data\01_primary_adr_detection\external_val_cadec\cadec_harmonised.csv")

    df_psytar = harmonise_psytar(psytar_excel, psytar_out)
    df_cadec, cadec_alignment_notes = harmonise_cadec(cadec_folder, cadec_out)
    print_st1_report(df_psytar, df_cadec, cadec_alignment_notes)


if __name__ == "__main__":
    main()
